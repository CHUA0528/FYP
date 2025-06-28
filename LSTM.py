import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
import pandas as pd
import numpy as np
import keras
from keras import layers, Sequential,losses,models,metrics,callbacks
import matplotlib.pyplot as plt
from sklearn.metrics import mean_absolute_error,root_mean_squared_error,r2_score,mean_absolute_percentage_error
import tensorflow as tf
from keras.optimizers import Adam
from openpyxl import load_workbook
from sklearn.preprocessing import MinMaxScaler

from sklearn.model_selection import train_test_split

model_file='C:/FYP/Dataset_Model/LSTM(Power)_1.keras'
dataset='Dataset/Data UTeM Electric Consumption.xlsx'

def timeseries_generator(dftest, window_size):
    q = []
    z = []
    for i in range(len(dftest) - window_size):
        row = [[a] for a in dftest.iloc[i:i + window_size]]
        label = dftest.iloc[i + window_size]
        q.append(row)

        z.append(label)


    x_train, x_test, y_train, y_test = train_test_split(q, z, random_state=0, shuffle=False, train_size=.9)

    return tf.convert_to_tensor(x_train, dtype=tf.float32), tf.convert_to_tensor(x_test, dtype=tf.float32),tf.convert_to_tensor (y_train,dtype=tf.float32),tf.convert_to_tensor(y_test,dtype=tf.float32)

def validation_result(history):
    min_val_epoch=np.argmin(history.history['val_loss'])+1
    min_val_loss=np.min(history.history['val_loss'])
    trained_epoch=len(history.history['val_loss'])
    plt.plot(history.history['loss'],label='train_loss')
    plt.plot(history.history['val_loss'],label='val_loss')
    plt.title(model.name + ' loss')
    plt.ylabel('loss')
    plt.xlabel('epoch')
    plt.scatter(np.argmin(history.history['val_loss']), min_val_loss, color='red',
                label=f'Min Val Loss({min_val_loss:.4f} at epoch {min_val_epoch})')
    plt.legend(loc='upper left',title=f"Trained for {trained_epoch} epochs")
    plt.show()

def load_model(ytest,xtest,file,scaler):
    model1 = models.load_model(file)


    prediction = scaler.inverse_transform(model1.predict(xtest))
    ytest = scaler.inverse_transform(np.reshape(ytest, (-1, 1)))

    mape_error = mean_absolute_percentage_error(ytest, prediction)
    rmse = root_mean_squared_error(ytest, prediction)
    mae_error = mean_absolute_error(ytest, prediction)
    r2 = r2_score(ytest, prediction)

    print(f'Root Mean Squared Error (RMSE): {rmse}')
    print(f'Mean Absolute Error (MAE): {mae_error}')
    print(f'Mean Absolute Percentage Error (MAPE): {mape_error}')
    print(f'R-squared (R²): {r2}')


    train_results = pd.DataFrame(data={'Test Predictions': prediction.flatten(), 'Actual': ytest.flatten()})
    print(train_results)

    # Add error metrics to the DataFrame (optional, one row for the errors)
    data_metrics = pd.DataFrame(data={
        'Metric': ['RMSE', 'MAE','MAPE', 'R²'],
        'Value': [rmse, mae_error,mape_error, r2]
    })

    # Concatenate the metrics with the predictions and actuals
    result_df = pd.concat([train_results, data_metrics], axis=1)
    file_name='PowerDataset_Result.xlsx'
    sheet_name = 'LSTM'

    # Check if the file already exists
    try:
        book = load_workbook(file_name)

        # Ensure sheet exists
        if sheet_name not in book.sheetnames:
            book.create_sheet(sheet_name)
            book.save(file_name)

        sheet = book[sheet_name]

        # === 🔁 Determine which run this is (1st, 2nd, 3rd...) ===
        num_columns = sheet.max_column
        block_width = 4  # Because each result has 4 columns: Prediction, Actual, Metric, Value
        run_index = num_columns // block_width + 1  # So run 1 = col 0, run 2 = col 4, etc.

        # === 🏷️ Rename result_df columns for this run ===
        result_df.columns = [
            f'Prediction {run_index}',
            f'Actual {run_index}',
            f'Metric {run_index}',
            f'Value {run_index}'
        ]

        # === ✅ Append the result to the right (side-by-side) ===
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            result_df.to_excel(writer,
                               sheet_name=sheet_name,
                               index=False,
                               header=True,
                               startrow=0,
                               startcol=(run_index - 1) * block_width)

    except FileNotFoundError:
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Results saved to 'model_results.xlsx'")


    plt.plot(train_results['Test Predictions'], label='Predictions')
    plt.plot(train_results['Actual'], label='Actual')
    plt.ylabel('Power (kW)')# Replace with the unit you're using.
    plt.xlabel('Timestep (per half-hourly)')  # Replace with the unit you're using.
    plt.grid()
    plt.xticks(np.arange(0, len(train_results)+10,10), rotation=45)
    plt.yticks(np.arange(400,np.max(ytest.flatten())+200,50))

    plt.title('Power Consumption versus Timestep')

    plt.legend()
    plt.show()


###Data Preprocessing
df_sheet1 = pd.read_excel(dataset,index_col='Date and Time', parse_dates=True,sheet_name='Oct 24')
df_sheet2 = pd.read_excel(dataset,index_col='Date and Time',parse_dates=True,sheet_name='Nov 24')
df_sheet3 = pd.read_excel(dataset,index_col='Date and Time',parse_dates=True,sheet_name='Dec 24')
df=pd.concat([df_sheet1,df_sheet2,df_sheet3])


df.index = pd.to_datetime(df.index)  # Ensure index is datetime
df = df.asfreq('30min')  # Set the frequency to 30 minutes





y_scaler =MinMaxScaler(feature_range=(0,1))
df['Energy Meter Reading, kWh'] = y_scaler.fit_transform(df[['Energy Meter Reading, kWh']])

WindowSize=32
X_train, X_test, Y_train, Y_test=timeseries_generator(df['Energy Meter Reading, kWh'],WindowSize)




##Model Building
model=Sequential([
    keras.Input(shape=(None,1)),
    layers.LSTM(128),
    layers.Dropout(0.2),
    layers.Dense(64, activation="relu"),  # Fully connected layer for feature extraction
    layers.Dropout(0.2),
    layers.Dense(1, activation="linear")  # Final output (single value prediction
])


###Model Compilation

model.compile(optimizer=Adam(learning_rate=0.001), loss=losses.MeanSquaredError(name='loss'),metrics=[keras.metrics.RootMeanSquaredError(name='rmse', dtype=None),'mae','mape'])

###Model Summary
model.summary()


###Model Callback Function
es = callbacks.EarlyStopping(monitor='val_loss',patience=3,mode='min')
cp1 = callbacks.ModelCheckpoint(model_file,monitor='val_loss', save_best_only=True)

###Model Training & Validation
result=model.fit(X_train,Y_train,validation_split=0.1,epochs=100,callbacks=[cp1,es])

###Model Validation Result
validation_result(result)

###Model Testing
load_model(Y_test,X_test,model_file,y_scaler)

